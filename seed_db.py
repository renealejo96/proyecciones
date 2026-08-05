import os
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import (
    Boolean, Column, DateTime, Float, Index, Integer, String, Text, create_engine
)
from sqlalchemy.orm import declarative_base, sessionmaker

DATABASE_URL = os.environ.get(
    "DATABASE_URL", "postgresql://postgres:admin@localhost:5432/proyecciones"
)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


class TpsrRecord(Base):
    __tablename__ = "tpsr_records"

    id = Column(Integer, primary_key=True, autoincrement=True)
    year = Column(Integer, nullable=True)
    source_week = Column(Integer, nullable=False, index=True)
    source_week_short = Column(String(10), nullable=True)
    date_val = Column(String(50), nullable=True)
    pruning_number = Column(String(50), default="")
    activity = Column(String(50), nullable=False, index=True)
    product_master = Column(String(100), nullable=False)
    product_master_norm = Column(String(100), nullable=False, index=True)
    product = Column(String(100), nullable=True)
    variety = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False, index=True)
    color = Column(String(50), nullable=True)
    material_origin = Column(String(100), nullable=True)
    block = Column(String(50), nullable=False, index=True)
    block_norm = Column(String(50), nullable=False, index=True)
    bed_location = Column(String(100), default="")
    total_beds = Column(Float, nullable=True)
    plants = Column(Integer, nullable=False, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class CycleDefinitionDB(Base):
    __tablename__ = "cycle_definitions"

    id = Column(Integer, primary_key=True, autoincrement=True)
    product_master = Column(String(100), nullable=False)
    product_master_norm = Column(String(100), nullable=False, index=True)
    variety = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False, index=True)
    activity = Column(String(50), nullable=False, index=True)
    cycle_weeks = Column(Integer, nullable=False, default=0)
    waste_rate = Column(Float, nullable=False, default=0.0)
    stems_per_plant = Column(Float, nullable=False, default=0.0)
    curve = Column(Text, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class RowAdjustmentDB(Base):
    __tablename__ = "row_adjustments"

    id = Column(Integer, primary_key=True, autoincrement=True)
    activity = Column(String(50), nullable=False)
    product_master_norm = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False)
    block = Column(String(50), nullable=False)
    block_norm = Column(String(50), nullable=False)
    source_week = Column(Integer, nullable=False)
    cycle_weeks = Column(Integer, nullable=True)
    waste_rate = Column(Float, nullable=True)
    stems_per_plant = Column(Float, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class MassAdjustmentDB(Base):
    __tablename__ = "mass_adjustments"

    id = Column(Integer, primary_key=True, autoincrement=True)
    product_master_norm = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False)
    activity = Column(String(50), nullable=False)
    cycle_weeks = Column(Integer, nullable=True)
    waste_rate = Column(Float, nullable=True)
    stems_per_plant = Column(Float, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class WeekAdjustmentDB(Base):
    __tablename__ = "week_adjustments"

    id = Column(Integer, primary_key=True, autoincrement=True)
    activity = Column(String(50), nullable=False)
    product_master_norm = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False)
    block = Column(String(50), nullable=False)
    block_norm = Column(String(50), nullable=False)
    source_week = Column(Integer, nullable=False)
    harvest_week = Column(Integer, nullable=False)
    agronomo_estimate = Column(Integer, nullable=True)
    real_closed = Column(Integer, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class BlockClosureDB(Base):
    __tablename__ = "block_closures"

    id = Column(Integer, primary_key=True, autoincrement=True)
    activity = Column(String(50), nullable=False)
    product_master_norm = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False)
    block = Column(String(50), nullable=False)
    block_norm = Column(String(50), nullable=False)
    source_week = Column(Integer, nullable=False)
    is_closed = Column(Boolean, default=False)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = text.upper().strip()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_week_code(raw_value: Any) -> int | None:
    if raw_value is None:
        return None
    digits = re.sub(r"\D", "", str(raw_value).strip())
    if not digits:
        return None
    if len(digits) == 4:
        year = 2000 + int(digits[:2])
        week = int(digits[2:])
    elif len(digits) == 6:
        year = int(digits[:4])
        week = int(digits[4:])
    else:
        return None
    return year * 100 + week


def format_short_week(week_code: int | None) -> str:
    if week_code is None:
        return ""
    year = week_code // 100
    week = week_code % 100
    return f"{year % 100:02d}{week:02d}"


def float_or_zero(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except ValueError:
        return 0.0


def init_db():
    Base.metadata.create_all(bind=engine)
    print("PostgreSQL tables created successfully.")


def seed_data_from_excel(workbook_path: Path):
    if not workbook_path.exists():
        print(f"File {workbook_path} does not exist. Skipping seed.")
        return

    db = SessionLocal()
    try:
        # Check if TpsrRecord table is already seeded
        existing_tpsr = db.query(TpsrRecord).first()
        if existing_tpsr:
            print("PostgreSQL database already has TPSR data. Skipping seed.")
            return

        print(f"Reading Excel {workbook_path} to seed PostgreSQL database...")
        t0 = time.time()
        wb = load_workbook(workbook_path, data_only=True)

        # 1. Seed CICLOS
        if "CICLOS" in wb.sheetnames:
            ws = wb["CICLOS"]
            cycles_to_add = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0] or not row[1] or not row[2]:
                    continue
                product_master = str(row[0]).strip()
                variety = str(row[1]).strip()
                activity = normalize_text(row[2])
                cycle_weeks = int(float_or_zero(row[3]))
                waste_rate = float_or_zero(row[4])
                stems_per_plant = float_or_zero(row[5])
                curve_vals = [str(float(c)) for c in row[6:] if c not in (None, "") and float_or_zero(c) > 0]
                curve_str = ",".join(curve_vals) if curve_vals else "1.0"

                cycles_to_add.append(
                    CycleDefinitionDB(
                        product_master=product_master,
                        product_master_norm=normalize_text(product_master),
                        variety=variety,
                        variety_norm=normalize_text(variety),
                        activity=activity,
                        cycle_weeks=cycle_weeks,
                        waste_rate=waste_rate,
                        stems_per_plant=stems_per_plant,
                        curve=curve_str,
                    )
                )
            db.bulk_save_objects(cycles_to_add)
            db.commit()
            print(f"Seeded {len(cycles_to_add)} cycle definitions.")

        # 2. Seed AJUSTES_CAMPO
        if "AJUSTES_CAMPO" in wb.sheetnames:
            ws = wb["AJUSTES_CAMPO"]
            row_adj_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                act = normalize_text(row[0])
                pm_norm = normalize_text(row[1])
                v_norm = normalize_text(row[2])
                blk = str(row[3] or "").strip()
                blk_norm = normalize_text(blk)
                sw = int(float_or_zero(row[4]))
                cw = int(row[5]) if row[5] not in (None, "") else None
                wr = float(row[6]) if row[6] not in (None, "") else None
                spp = float(row[7]) if row[7] not in (None, "") else None
                row_adj_list.append(
                    RowAdjustmentDB(
                        activity=act,
                        product_master_norm=pm_norm,
                        variety_norm=v_norm,
                        block=blk,
                        block_norm=blk_norm,
                        source_week=sw,
                        cycle_weeks=cw,
                        waste_rate=wr,
                        stems_per_plant=spp,
                    )
                )
            db.bulk_save_objects(row_adj_list)
            db.commit()
            print(f"Seeded {len(row_adj_list)} row adjustments.")

        # 3. Seed AJUSTES_MASIVOS
        if "AJUSTES_MASIVOS" in wb.sheetnames:
            ws = wb["AJUSTES_MASIVOS"]
            mass_adj_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                pm_norm = normalize_text(row[0]) if row[0] != "*" else "*"
                v_norm = normalize_text(row[1]) if row[1] != "*" else "*"
                act = normalize_text(row[2]) if row[2] != "*" else "*"
                cw = int(row[3]) if row[3] not in (None, "") else None
                wr = float(row[4]) if row[4] not in (None, "") else None
                spp = float(row[5]) if row[5] not in (None, "") else None
                mass_adj_list.append(
                    MassAdjustmentDB(
                        product_master_norm=pm_norm,
                        variety_norm=v_norm,
                        activity=act,
                        cycle_weeks=cw,
                        waste_rate=wr,
                        stems_per_plant=spp,
                    )
                )
            db.bulk_save_objects(mass_adj_list)
            db.commit()
            print(f"Seeded {len(mass_adj_list)} mass adjustments.")

        # 4. Seed AJUSTES_SEMANA
        if "AJUSTES_SEMANA" in wb.sheetnames:
            ws = wb["AJUSTES_SEMANA"]
            week_adj_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                act = normalize_text(row[0])
                pm_norm = normalize_text(row[1])
                v_norm = normalize_text(row[2])
                blk = str(row[3] or "").strip()
                blk_norm = normalize_text(blk)
                sw = int(float_or_zero(row[4]))
                hw = int(float_or_zero(row[5]))
                ae = int(row[6]) if row[6] not in (None, "") else None
                rc = int(row[7]) if row[7] not in (None, "") else None
                week_adj_list.append(
                    WeekAdjustmentDB(
                        activity=act,
                        product_master_norm=pm_norm,
                        variety_norm=v_norm,
                        block=blk,
                        block_norm=blk_norm,
                        source_week=sw,
                        harvest_week=hw,
                        agronomo_estimate=ae,
                        real_closed=rc,
                    )
                )
            db.bulk_save_objects(week_adj_list)
            db.commit()
            print(f"Seeded {len(week_adj_list)} week adjustments.")

        # 5. Seed CIERRE_BLOQUES (if present)
        if "CIERRE_BLOQUES" in wb.sheetnames:
            ws = wb["CIERRE_BLOQUES"]
            closures_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                act = normalize_text(row[0])
                pm_norm = normalize_text(row[1])
                v_norm = normalize_text(row[2])
                blk = str(row[3] or "").strip()
                blk_norm = normalize_text(blk)
                sw = int(float_or_zero(row[4]))
                ic = bool(row[5])
                closures_list.append(
                    BlockClosureDB(
                        activity=act,
                        product_master_norm=pm_norm,
                        variety_norm=v_norm,
                        block=blk,
                        block_norm=blk_norm,
                        source_week=sw,
                        is_closed=ic,
                    )
                )
            db.bulk_save_objects(closures_list)
            db.commit()
            print(f"Seeded {len(closures_list)} block closures.")

        # 6. Seed TPSR
        if "TPSR" in wb.sheetnames:
            ws = wb["TPSR"]
            tpsr_records = []
            for r in range(2, ws.max_row + 1):
                act = normalize_text(ws.cell(r, 6).value)
                if act not in {"SIEMBRA", "PODA"}:
                    continue
                sw = parse_week_code(ws.cell(r, 2).value)
                if sw is None:
                    continue

                pm = str(ws.cell(r, 7).value or "").strip()
                variety = str(ws.cell(r, 9).value or "").strip()
                block = str(ws.cell(r, 12).value or "").strip()
                plants = int(round(float_or_zero(ws.cell(r, 15).value)))

                tpsr_records.append(
                    TpsrRecord(
                        year=int(ws.cell(r, 1).value) if ws.cell(r, 1).value and str(ws.cell(r, 1).value).isdigit() else None,
                        source_week=sw,
                        source_week_short=format_short_week(sw),
                        date_val=str(ws.cell(r, 4).value or "").strip(),
                        pruning_number=str(ws.cell(r, 5).value or "").strip(),
                        activity=act,
                        product_master=pm,
                        product_master_norm=normalize_text(pm),
                        product=str(ws.cell(r, 8).value or "").strip(),
                        variety=variety,
                        variety_norm=normalize_text(variety),
                        color=str(ws.cell(r, 10).value or "").strip(),
                        material_origin=str(ws.cell(r, 11).value or "").strip(),
                        block=block,
                        block_norm=normalize_text(block),
                        bed_location=str(ws.cell(r, 13).value or "").strip(),
                        total_beds=float_or_zero(ws.cell(r, 14).value),
                        plants=plants,
                    )
                )

            # Insert in chunks of 2000 for speed
            chunk_size = 2000
            for i in range(0, len(tpsr_records), chunk_size):
                db.bulk_save_objects(tpsr_records[i : i + chunk_size])
                db.commit()
            print(f"Seeded {len(tpsr_records)} TPSR records in {round(time.time() - t0, 2)}s.")

    finally:
        db.close()


if __name__ == "__main__":
    init_db()
    seed_data_from_excel(Path("siembras_podas/tpsr.xlsx"))
