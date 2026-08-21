import os
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import (
    Boolean, Column, DateTime, Float, Index, Integer, String, Text, create_engine, text
)
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import declarative_base, sessionmaker

DATABASE_URL = os.environ.get(
    "DATABASE_URL", "postgresql://postgres:admin@localhost:5432/proyecciones"
)

# Connect to database with pool resilience
engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


class TpsrRecord(Base):
    __tablename__ = "tpsr_records"

    id = Column(Integer, primary_key=True, autoincrement=True)
    year = Column(Integer, nullable=True)
    source_week = Column(Integer, nullable=False, index=True)
    source_week_short = Column(String(10), nullable=True)
    date_val = Column(String(50), nullable=True)
    pruning_number = Column(String(50), default="")
    activity = Column(String(50), nullable=False, index=True)  # 'SIEMBRA' or 'PODA'
    product_master = Column(String(100), nullable=False)
    product_master_norm = Column(String(100), nullable=False, index=True)
    product = Column(String(100), nullable=True)
    variety = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False, index=True)
    color = Column(String(50), nullable=True)
    material_origin = Column(String(100), nullable=True)
    block = Column(String(50), nullable=False, index=True)
    block_norm = Column(String(50), nullable=False, index=True)
    bed_location = Column(String(100), default="")
    total_beds = Column(Float, nullable=True)
    plants = Column(Integer, nullable=False, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class CycleDefinitionDB(Base):
    __tablename__ = "cycle_definitions"

    id = Column(Integer, primary_key=True, autoincrement=True)
    product_master = Column(String(100), nullable=False)
    product_master_norm = Column(String(100), nullable=False, index=True)
    variety = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False, index=True)
    activity = Column(String(50), nullable=False, index=True)
    cycle_weeks = Column(Integer, nullable=False, default=0)
    waste_rate = Column(Float, nullable=False, default=0.0)
    stems_per_plant = Column(Float, nullable=False, default=0.0)
    curve = Column(Text, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class RowAdjustmentDB(Base):
    __tablename__ = "row_adjustments"

    id = Column(Integer, primary_key=True, autoincrement=True)
    activity = Column(String(50), nullable=False)
    product_master_norm = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False)
    block = Column(String(50), nullable=False)
    block_norm = Column(String(50), nullable=False)
    source_week = Column(Integer, nullable=False)
    cycle_weeks = Column(Integer, nullable=True)
    waste_rate = Column(Float, nullable=True)
    stems_per_plant = Column(Float, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class MassAdjustmentDB(Base):
    __tablename__ = "mass_adjustments"

    id = Column(Integer, primary_key=True, autoincrement=True)
    product_master_norm = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False)
    activity = Column(String(50), nullable=False)
    cycle_weeks = Column(Integer, nullable=True)
    waste_rate = Column(Float, nullable=True)
    stems_per_plant = Column(Float, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class WeekAdjustmentDB(Base):
    __tablename__ = "week_adjustments"

    id = Column(Integer, primary_key=True, autoincrement=True)
    activity = Column(String(50), nullable=False)
    product_master_norm = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False)
    block = Column(String(50), nullable=False)
    block_norm = Column(String(50), nullable=False)
    source_week = Column(Integer, nullable=False)
    harvest_week = Column(Integer, nullable=False)
    agronomo_estimate = Column(Integer, nullable=True)
    real_closed = Column(Integer, nullable=True)
    dump_stems = Column(Integer, default=0, nullable=True)
    is_dump = Column(Boolean, default=False, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class BlockClosureDB(Base):
    __tablename__ = "block_closures"

    id = Column(Integer, primary_key=True, autoincrement=True)
    activity = Column(String(50), nullable=False)
    product_master_norm = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False)
    block = Column(String(50), nullable=False)
    block_norm = Column(String(50), nullable=False)
    source_week = Column(Integer, nullable=False)
    is_closed = Column(Boolean, default=False)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class EstimateAuditDB(Base):
    __tablename__ = "estimate_audits"

    id = Column(Integer, primary_key=True, autoincrement=True)
    product_master = Column(String(100), nullable=False)
    product_master_norm = Column(String(100), nullable=False, index=True)
    variety = Column(String(100), nullable=False)
    variety_norm = Column(String(100), nullable=False, index=True)
    block = Column(String(50), nullable=False, index=True)
    block_norm = Column(String(50), nullable=False, index=True)
    activity = Column(String(50), nullable=False, index=True)
    source_week = Column(Integer, nullable=False, index=True)
    harvest_week = Column(Integer, nullable=False, index=True)
    mode = Column(String(50), default="AGRONOMO")  # 'AGRONOMO' or 'REAL'
    action = Column(String(50), default="EDIT")    # 'EDIT', 'RESTORE', 'RESET', 'DELETE', 'INITIAL'
    previous_value = Column(Integer, nullable=True)
    new_value = Column(Integer, nullable=True)
    dump_stems = Column(Integer, nullable=True, default=0)
    notes = Column(String(255), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, index=True)


class AppMetaDB(Base):
    __tablename__ = "app_meta"

    key = Column(String(50), primary_key=True)
    value = Column(String(255), nullable=False)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


def get_data_version(db_session=None) -> int:
    close_session = False
    if db_session is None:
        db_session = SessionLocal()
        close_session = True
    try:
        meta = db_session.query(AppMetaDB).filter_by(key="data_version").first()
        if meta and meta.value:
            return int(meta.value)
        return 0
    except Exception:
        return 0
    finally:
        if close_session:
            db_session.close()


def bump_data_version(db_session=None) -> int:
    close_session = False
    if db_session is None:
        db_session = SessionLocal()
        close_session = True
    try:
        meta = db_session.query(AppMetaDB).filter_by(key="data_version").first()
        if not meta:
            meta = AppMetaDB(key="data_version", value="1")
            db_session.add(meta)
        else:
            new_val = int(meta.value or 0) + 1
            meta.value = str(new_val)
            meta.updated_at = datetime.utcnow()
        db_session.commit()
        return int(meta.value)
    except Exception:
        db_session.rollback()
        return int(time.time())
    finally:
        if close_session:
            db_session.close()


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = text.upper().strip()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_week_code(raw_value: Any) -> int | None:
    if raw_value is None:
        return None
    digits = re.sub(r"\D", "", str(raw_value).strip())
    if not digits:
        return None
    if len(digits) == 4:
        year = 2000 + int(digits[:2])
        week = int(digits[2:])
    elif len(digits) == 6:
        year = int(digits[:4])
        week = int(digits[4:])
    else:
        return None
    return year * 100 + week


def format_short_week(week_code: int | None) -> str:
    if week_code is None:
        return ""
    year = week_code // 100
    week = week_code % 100
    return f"{year % 100:02d}{week:02d}"


def float_or_zero(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except ValueError:
        return 0.0


def init_db():
    Base.metadata.create_all(bind=engine)
    # Automatic migration: ensure is_dump and dump_stems columns exist in week_adjustments
    with engine.connect() as conn:
        try:
            conn.execute(text("ALTER TABLE week_adjustments ADD COLUMN IF NOT EXISTS is_dump BOOLEAN DEFAULT FALSE;"))
            conn.execute(text("ALTER TABLE week_adjustments ADD COLUMN IF NOT EXISTS dump_stems INTEGER DEFAULT 0;"))
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            try:
                conn.execute(text("ALTER TABLE week_adjustments ADD COLUMN is_dump BOOLEAN DEFAULT FALSE;"))
                conn.commit()
            except Exception:
                pass
            try:
                conn.execute(text("ALTER TABLE week_adjustments ADD COLUMN dump_stems INTEGER DEFAULT 0;"))
                conn.commit()
            except Exception:
                pass

    db = SessionLocal()
    try:
        meta = db.query(AppMetaDB).filter_by(key="data_version").first()
        if not meta:
            meta = AppMetaDB(key="data_version", value="1")
            db.add(meta)
            db.commit()

        # Seed baseline audit history from current adjustments if table is empty
        audit_count = db.query(EstimateAuditDB).count()
        if audit_count == 0:
            existing_adjs = db.query(WeekAdjustmentDB).all()
            for a in existing_adjs:
                val = a.agronomo_estimate if a.agronomo_estimate is not None else a.real_closed
                if val is not None or a.is_dump or a.dump_stems:
                    db.add(
                        EstimateAuditDB(
                            product_master=a.product_master_norm,
                            product_master_norm=a.product_master_norm,
                            variety=a.variety_norm,
                            variety_norm=a.variety_norm,
                            block=a.block,
                            block_norm=a.block_norm,
                            activity=a.activity,
                            source_week=a.source_week,
                            harvest_week=a.harvest_week,
                            mode="AGRONOMO" if a.agronomo_estimate is not None else "REAL",
                            action="INITIAL",
                            previous_value=None,
                            new_value=val,
                            dump_stems=a.dump_stems or 0,
                            notes="Registro inicial",
                            created_at=a.updated_at or datetime.utcnow(),
                        )
                    )
            db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


def seed_data_from_excel_if_empty(workbook_path: Path):
    init_db()
    db = SessionLocal()
    try:
        count = db.query(TpsrRecord).count()
        if count > 0:
            return  # Already seeded
        from seed_db import seed_data_from_excel
        seed_data_from_excel(workbook_path)
    finally:
        db.close()


def process_tpsr_excel_upload(file_input: Any, commit: bool = True) -> Dict[str, Any]:
    """
    Reads an uploaded TPSR Excel file (.xlsx), validates each row,
    performs deduplication against PostgreSQL, detects plant mismatches/corrections,
    and inserts/updates accordingly.
    """
    wb = load_workbook(file_input, data_only=True)
    sheet_name = "TPSR" if "TPSR" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    db = SessionLocal()
    try:
        # Load existing TPSR records for fast lookup
        existing_records = db.query(TpsrRecord).all()
        # Lookup dictionary: key -> TpsrRecord object
        # Natural key: (source_week, activity_norm, product_master_norm, variety_norm, block_norm, bed_location, pruning_number)
        lookup_dict: Dict[Tuple[int, str, str, str, str, str, str], TpsrRecord] = {}
        # Fallback dictionary without bed_location if bed_location is empty
        fallback_dict: Dict[Tuple[int, str, str, str, str], TpsrRecord] = {}

        for rec in existing_records:
            nat_key = (
                rec.source_week,
                rec.activity,
                rec.product_master_norm,
                rec.variety_norm,
                rec.block_norm,
                str(rec.bed_location or "").strip(),
                str(rec.pruning_number or "").strip(),
            )
            lookup_dict[nat_key] = rec
            fallback_key = (
                rec.source_week,
                rec.activity,
                rec.product_master_norm,
                rec.variety_norm,
                rec.block_norm,
            )
            fallback_dict[fallback_key] = rec

        new_count = 0
        updated_count = 0
        skipped_count = 0
        invalid_count = 0

        new_records_list: List[TpsrRecord] = []
        plant_adjustments_details: List[Dict[str, Any]] = []

        for r in range(2, ws.max_row + 1):
            act = normalize_text(ws.cell(r, 6).value)
            if act not in {"SIEMBRA", "PODA"}:
                invalid_count += 1
                continue
            sw = parse_week_code(ws.cell(r, 2).value)
            if sw is None:
                invalid_count += 1
                continue

            pm = str(ws.cell(r, 7).value or "").strip()
            variety = str(ws.cell(r, 9).value or "").strip()
            block = str(ws.cell(r, 12).value or "").strip()
            bed_location = str(ws.cell(r, 13).value or "").strip()
            pruning_number = str(ws.cell(r, 5).value or "").strip()
            plants = int(round(float_or_zero(ws.cell(r, 15).value)))

            if not pm or not variety or not block:
                invalid_count += 1
                continue

            if "VERONICA" in pm.upper() or "SPLASH" in normalize_text(variety):
                pm = "VERONICA"
                pm_norm = "VERONICA"
            else:
                pm_norm = normalize_text(pm)

            v_norm = normalize_text(variety)
            b_norm = normalize_text(block)

            nat_key = (sw, act, pm_norm, v_norm, b_norm, bed_location, pruning_number)
            fallback_key = (sw, act, pm_norm, v_norm, b_norm)

            existing_rec = lookup_dict.get(nat_key) or fallback_dict.get(fallback_key)

            if existing_rec:
                # Deduplication check
                if existing_rec.plants == plants:
                    skipped_count += 1
                else:
                    # Mismatch detected in plant reporting! (Correction)
                    old_plants = existing_rec.plants
                    if commit:
                        existing_rec.plants = plants
                        existing_rec.updated_at = datetime.utcnow()
                    updated_count += 1
                    plant_adjustments_details.append({
                        "id": existing_rec.id,
                        "activity": act,
                        "product_master": pm,
                        "variety": variety,
                        "block": block,
                        "source_week": sw,
                        "source_week_short": format_short_week(sw),
                        "old_plants": old_plants,
                        "new_plants": plants,
                        "difference": plants - old_plants,
                    })
            else:
                # Brand new record
                new_rec = TpsrRecord(
                    year=int(ws.cell(r, 1).value) if ws.cell(r, 1).value and str(ws.cell(r, 1).value).isdigit() else None,
                    source_week=sw,
                    source_week_short=format_short_week(sw),
                    date_val=str(ws.cell(r, 4).value or "").strip(),
                    pruning_number=pruning_number,
                    activity=act,
                    product_master=pm,
                    product_master_norm=pm_norm,
                    product=str(ws.cell(r, 8).value or "").strip(),
                    variety=variety,
                    variety_norm=v_norm,
                    color=str(ws.cell(r, 10).value or "").strip(),
                    material_origin=str(ws.cell(r, 11).value or "").strip(),
                    block=block,
                    block_norm=b_norm,
                    bed_location=bed_location,
                    total_beds=float_or_zero(ws.cell(r, 14).value),
                    plants=plants,
                )
                new_records_list.append(new_rec)
                lookup_dict[nat_key] = new_rec
                new_count += 1

        if commit:
            if new_records_list:
                db.bulk_save_objects(new_records_list)
            db.commit()

        return {
            "ok": True,
            "new_count": new_count,
            "updated_count": updated_count,
            "skipped_count": skipped_count,
            "invalid_count": invalid_count,
            "total_processed": new_count + updated_count + skipped_count,
            "plant_adjustments": plant_adjustments_details,
            "committed": commit,
        }
    except Exception as exc:
        db.rollback()
        return {"ok": False, "message": str(exc)}
    finally:
        db.close()
